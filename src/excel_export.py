#!/usr/bin/env python3
"""
Excel export helper using openpyxl.
Called from C++ via QProcess with JSON input.
"""

import sys
import json
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, numbers
except ImportError:
    print("ERROR: openpyxl not installed. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def create_daily_report(data: dict, output_path: str) -> bool:
    """
    Create an Excel daily report from the provided dashboard data.

    Args:
        data: Dictionary with keys: date, today, mtd, mtd_ly, cloudbeds
        output_path: Path where to save the Excel file

    Returns:
        True if successful, False otherwise
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Daily Report"

        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15

        # Formats
        title_font = Font(name='Calibri', size=14, bold=True)
        bold_font = Font(bold=True)
        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        currency_fmt = '_("$"* #,##0.00_);_("$"* (#,##0.00);_("$"* "-"??_);_(@_)'
        percent_fmt = '0.0%'

        # Title
        ws['A1'] = "Financial Dashboard Report"
        ws['A1'].font = title_font

        # Date
        ws['A2'] = "Date:"
        ws['A2'].font = bold_font
        ws['B2'] = data.get('date', '')

        # Headers (row 4)
        headers = ['Revenue Category', 'Today', 'MTD', 'MTD-LY', 'Variance $', 'Variance %']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = bold_font
            cell.fill = header_fill

        # LSK Revenue Groups (rows 5-10)
        lsk_groups = ['Cafe Bar', 'Cafe Food', 'Health Club', 'Laundry', 'Misc', 'Retail']
        today_data = data.get('today', {})
        mtd_data = data.get('mtd', {})
        mtd_ly_data = data.get('mtd_ly', {})

        for row_offset, group_name in enumerate(lsk_groups):
            row = 5 + row_offset
            ws.cell(row=row, column=1).value = group_name

            # Today
            today_val = today_data.get(group_name, 0)
            today_cell = ws.cell(row=row, column=2)
            today_cell.value = today_val
            today_cell.number_format = currency_fmt

            # MTD
            mtd_val = mtd_data.get(group_name, 0)
            mtd_cell = ws.cell(row=row, column=3)
            mtd_cell.value = mtd_val
            mtd_cell.number_format = currency_fmt

            # MTD-LY
            mtd_ly_val = mtd_ly_data.get(group_name, 0)
            mtd_ly_cell = ws.cell(row=row, column=4)
            mtd_ly_cell.value = mtd_ly_val
            mtd_ly_cell.number_format = currency_fmt

            # Variance $ (MTD - MTD-LY)
            var_cell = ws.cell(row=row, column=5)
            var_cell.value = f'=C{row}-D{row}'
            var_cell.number_format = currency_fmt

            # Variance % (IF(MTD-LY=0, 0%, (MTD-MTD-LY)/MTD-LY))
            var_pct_cell = ws.cell(row=row, column=6)
            var_pct_cell.value = f'=IF(D{row}=0,0,(C{row}-D{row})/D{row})'
            var_pct_cell.number_format = percent_fmt

        # Cloudbeds data
        cb_data = data.get('cloudbeds', {})
        if cb_data:
            row = 12
            cloudbeds_labels = [
                ('Room Revenue (Today):', cb_data.get('room_revenue', 0), True),
                ('Occupied Rooms:', cb_data.get('occupied_rooms', 0), False),
                ('Total Rooms:', cb_data.get('total_rooms', 0), False),
                ('ADR (Today):', cb_data.get('adr', 0), True),
                ('RevPAR (Today):', cb_data.get('revpar', 0), True),
                ('MTD Revenue:', cb_data.get('mtd_revenue', 0), True),
            ]

            for label, value, is_currency in cloudbeds_labels:
                label_cell = ws.cell(row=row, column=1)
                label_cell.value = label
                label_cell.font = bold_font

                value_cell = ws.cell(row=row, column=2)
                value_cell.value = value
                if is_currency:
                    value_cell.number_format = currency_fmt

                row += 1

        # Save workbook
        wb.save(output_path)
        print(f"Successfully saved: {output_path}")
        return True

    except Exception as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return False


if __name__ == '__main__':
    if len(sys.argv) != 3:
        print("Usage: python excel_export.py <json_input> <output_path>", file=sys.stderr)
        sys.exit(1)

    json_input = sys.argv[1]
    output_path = sys.argv[2]

    try:
        # Parse JSON input
        data = json.loads(json_input)
    except json.JSONDecodeError as e:
        print(f"ERROR: Invalid JSON input: {e}", file=sys.stderr)
        sys.exit(1)

    # Create the report
    success = create_daily_report(data, output_path)
    sys.exit(0 if success else 1)
